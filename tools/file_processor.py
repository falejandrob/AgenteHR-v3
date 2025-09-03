import os
import logging
from typing import Dict, List, Optional
import PyPDF2
import openpyxl
from werkzeug.utils import secure_filename

logger = logging.getLogger(__name__)

class FileProcessor:
    """Class to handle file upload and processing for PDF and XLSX files"""
    
    ALLOWED_EXTENSIONS = {'pdf', 'xlsx'}
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
    
    def __init__(self, upload_folder: str = "uploads"):
        self.upload_folder = upload_folder
        os.makedirs(upload_folder, exist_ok=True)
    
    def allowed_file(self, filename: str) -> bool:
        """Check if file extension is allowed"""
        return '.' in filename and \
               filename.rsplit('.', 1)[1].lower() in self.ALLOWED_EXTENSIONS
    
    def save_file(self, file, session_id: str) -> Optional[Dict]:
        """Save uploaded file and return file info"""
        try:
            if not file or file.filename == '':
                return None
            
            if not self.allowed_file(file.filename):
                raise ValueError(f"File type not allowed. Only {', '.join(self.ALLOWED_EXTENSIONS)} are supported.")
            
            # Check file size
            file.seek(0, os.SEEK_END)
            file_size = file.tell()
            file.seek(0)
            
            if file_size > self.MAX_FILE_SIZE:
                raise ValueError(f"File too large. Maximum size is {self.MAX_FILE_SIZE // (1024*1024)}MB")
            
            # Secure filename
            filename = secure_filename(file.filename)
            file_extension = filename.rsplit('.', 1)[1].lower()
            
            # Create unique filename with session_id
            unique_filename = f"{session_id}_{filename}"
            file_path = os.path.join(self.upload_folder, unique_filename)
            
            # Save file
            file.save(file_path)
            
            file_info = {
                'original_name': filename,
                'file_path': file_path,
                'file_type': file_extension,
                'file_size': file_size,
                'session_id': session_id
            }
            
            logger.info(f"File saved: {unique_filename} ({file_size} bytes)")
            return file_info
            
        except Exception as e:
            logger.error(f"Error saving file: {e}")
            raise
    
    def process_pdf(self, file_path: str) -> str:
        """Extract text from PDF file"""
        try:
            text_content = ""
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                
                for page_num, page in enumerate(pdf_reader.pages):
                    try:
                        page_text = page.extract_text()
                        if page_text.strip():
                            text_content += f"\n--- Page {page_num + 1} ---\n"
                            text_content += page_text.strip() + "\n"
                    except Exception as e:
                        logger.warning(f"Error extracting page {page_num + 1}: {e}")
                        continue
            
            if not text_content.strip():
                raise ValueError("No text could be extracted from the PDF")
            
            logger.info(f"PDF processed successfully: {len(text_content)} characters extracted")
            return text_content.strip()
            
        except Exception as e:
            logger.error(f"Error processing PDF {file_path}: {e}")
            raise
    
    def process_excel(self, file_path: str) -> str:
        """Extract text from Excel file"""
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            text_content = ""
            
            for sheet_name in workbook.sheetnames:
                try:
                    sheet = workbook[sheet_name]
                    text_content += f"\n--- Sheet: {sheet_name} ---\n"
                    
                    # Process rows
                    for row_num, row in enumerate(sheet.iter_rows(values_only=True), 1):
                        row_data = []
                        for cell in row:
                            if cell is not None:
                                cell_str = str(cell).strip()
                                if cell_str:
                                    row_data.append(cell_str)
                        
                        if row_data:
                            text_content += f"Row {row_num}: " + " | ".join(row_data) + "\n"
                    
                except Exception as e:
                    logger.warning(f"Error processing sheet {sheet_name}: {e}")
                    continue
            
            workbook.close()
            
            if not text_content.strip():
                raise ValueError("No data could be extracted from the Excel file")
            
            logger.info(f"Excel processed successfully: {len(text_content)} characters extracted")
            return text_content.strip()
            
        except Exception as e:
            logger.error(f"Error processing Excel {file_path}: {e}")
            raise
    
    def process_file(self, file_info: Dict) -> str:
        """Process file and extract text based on file type"""
        file_path = file_info['file_path']
        file_type = file_info['file_type']
        
        try:
            if file_type == 'pdf':
                return self.process_pdf(file_path)
            elif file_type == 'xlsx':
                return self.process_excel(file_path)
            else:
                raise ValueError(f"Unsupported file type: {file_type}")
                
        except Exception as e:
            logger.error(f"Error processing file {file_path}: {e}")
            raise
    
    def get_session_files(self, session_id: str) -> List[Dict]:
        """Get list of files for a session"""
        session_files = []
        
        try:
            if not os.path.exists(self.upload_folder):
                return session_files
            
            for filename in os.listdir(self.upload_folder):
                if filename.startswith(f"{session_id}_"):
                    file_path = os.path.join(self.upload_folder, filename)
                    if os.path.isfile(file_path):
                        original_name = filename[len(f"{session_id}_"):]
                        file_extension = original_name.rsplit('.', 1)[1].lower()
                        file_size = os.path.getsize(file_path)
                        
                        session_files.append({
                            'original_name': original_name,
                            'file_path': file_path,
                            'file_type': file_extension,
                            'file_size': file_size,
                            'session_id': session_id
                        })
        
        except Exception as e:
            logger.error(f"Error getting session files: {e}")
        
        return session_files
    
    def clear_session_files(self, session_id: str):
        """Remove all files for a session"""
        try:
            session_files = self.get_session_files(session_id)
            for file_info in session_files:
                file_path = file_info['file_path']
                if os.path.exists(file_path):
                    os.remove(file_path)
                    logger.info(f"Removed file: {file_path}")
        
        except Exception as e:
            logger.error(f"Error clearing session files: {e}")
    
    def cleanup_old_files(self, max_age_hours: int = 1) -> int:
        """Remove files older than max_age_hours"""
        deleted_count = 0
        
        try:
            if not os.path.exists(self.upload_folder):
                return deleted_count
            
            import time
            current_time = time.time()
            max_age_seconds = max_age_hours * 3600
            
            for filename in os.listdir(self.upload_folder):
                file_path = os.path.join(self.upload_folder, filename)
                
                if os.path.isfile(file_path):
                    file_age = current_time - os.path.getmtime(file_path)
                    
                    if file_age > max_age_seconds:
                        try:
                            os.remove(file_path)
                            deleted_count += 1
                            logger.info(f"Deleted old file: {filename}")
                        except Exception as e:
                            logger.error(f"Error deleting file {filename}: {e}")
            
            logger.info(f"Cleanup completed: {deleted_count} files deleted")
            
        except Exception as e:
            logger.error(f"Error during cleanup: {e}")
        
        return deleted_count
